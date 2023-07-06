# -*- coding: utf-8 -*-
"""
Created on Sat Apr 15 12:09:02 2023

@author: rm1221
"""


import glob
import tkinter as tk
from pathlib import Path
from tkinter import ttk
import time
import datetime
from PIL import Image, ImageTk
import cv2
import openpyxl
from openpyxl import Workbook

import pathlib
from skimage.metrics import structural_similarity
global current_image_id
global time
global Serialno
Serialno=1
current_image_id=1
time1=0


class App(tk.Tk):
    
    def __init__(self):
        super().__init__()

        self.title('Tkinter PhotoImage Demo')
       
        self.time = time1
        self.attributes('-fullscreen', True)
        self.configure(bg='black')
        self.base_width = 1230
        self.failedRoi=0
        self.serialno =Serialno
        self.current_image_id = current_image_id
        self.next_button = tk.Button(self, text="NEXT", font=("fixed", 30, "bold"), command=self.next_pic)
        self.timer = ttk.Label(self, text=self.time,font=("fixed", 30, "bold"))
        self.image_view = ttk.Label(self)
        self.image_id=ttk.Label(self, text=self.current_image_id,font=("fixed", 30, "bold"))
        self.fault=ttk.Label(self, text=self.failedRoi,font=("fixed", 30, "bold"))
        self.image_list = [Path(img) for img in glob.glob("images/*.png")]       
        self.max_image_id = len(self.image_list) 
        self.image_view.pack(side="top",expand=True)
        self.change_pic(self.image_list[self.current_image_id])
        self.next_button.pack(side="right")
        self.timer.pack(side="right",padx=150 )
        self.image_id.pack(side="left",padx=200)
        self.fault.pack(side="left",padx=150)
        self.exposure=-7
        self.camera_num=0
        self.after(1000, self.update)  # start the update 1 second later

    def change_pic(self, image_path):
        self.image = Image.open(Path(image_path))
        w, h = self.image.size
        new_height = int(self.base_width / w * h)
        self.image = self.image.resize((self.base_width, new_height))
        self.python_image = ImageTk.PhotoImage(self.image)
        self.image_view.configure(image=self.python_image)
        #self.image_view=self.image
        self.image_id.configure(text=self.current_image_id)
        if self.current_image_id == 1:
            self.time=15
        elif self.current_image_id == 2:
            self.time=10
        elif self.current_image_id == 3:
            self.time=10
        elif self.current_image_id == 4:
            self.time=10
        elif self.current_image_id == 5:
            self.time=5
        elif self.current_image_id == 6:
            self.time=10
        elif self.current_image_id == 7:
            self.time=65
        elif self.current_image_id == 8:
            self.time=15
            #self.timer.configure(text=self.time,state="disabled")
        self.timer.configure(text=self.time)
        file=openpyxl.load_workbook("data.xlsx")
        sheet=file.active
        sheet.cell(column = 4,row=sheet.max_row+1,value=self.time)
        file.save("data.xlsx")
        
        
        
    def stillimage(self):
        self.image1 = Image.open('blank/1.png')
        w, h = self.image1.size
        new_height = int(self.base_width / w * h)
        self.image1 = self.image1.resize((self.base_width, new_height))
        self.python_image1 = ImageTk.PhotoImage(self.image1)
        self.image_view.configure(image=self.python_image1)
        self.update_idletasks()
        
        

    def next_pic(self):
        now=datetime.datetime.now()
        file=openpyxl.load_workbook("data.xlsx")
        sheet=file.active
        sheet.cell(column = 1,row=sheet.max_row,value=self.serialno)
        sheet.cell(column = 2,row=sheet.max_row,value=self.current_image_id)
        sheet.cell(column = 3,row=sheet.max_row,value=now.strftime("%y-%m-%d %H:%M:%S"))
        sheet.cell(column = 5,row=sheet.max_row,value=self.time)
        file.save("data.xlsx")
        
        
        
        self.stillimage()
        self.video_capture()
        self.screw_det()
        if self.failedRoi==0:   
            self.current_image_id += 1
            if self.current_image_id > 8:
                self.current_image_id=1
                self.serialno+=1
            
            
            self.change_pic(self.image_list[self.current_image_id])
            self.image_id.configure(text=self.current_image_id)
            self.fault.configure(text=self.failedRoi)
            file=openpyxl.load_workbook("data.xlsx")
            sheet=file.active
           
            sheet.cell(column = 6,row=sheet.max_row-1,value="PASS")
            sheet.cell(column = 7,row=sheet.max_row-1,value=self.failedRoi)
            file.save("data.xlsx")
           
            
            
            return self.current_image_id 
        else:
            
            if self.current_image_id == 0:
                return
            self.current_image_id= self.current_image_id 
            self.change_pic(self.image_list[self.current_image_id])
            self.image_id.configure(text=self.current_image_id)
            self.fault.configure(text=self.failedRoi)
            file=openpyxl.load_workbook("data.xlsx")
            sheet=file.active
            sheet.cell(column = 6,row=sheet.max_row-1,value="FAIL")
            sheet.cell(column = 7,row=sheet.max_row-1,value=self.failedRoi)
            file.save("data.xlsx")
           
            
            return self.current_image_id
        
    def update(self):
        if (self.current_image_id <=7):
            self.time -= 1
            
        self.timer.configure(text=self.time)
        self.after(1000, self.update)
         
        
    def video_capture(self):
       
        i=0;
        self.cap=cv2.VideoCapture()
        ret,frame=self.cap.read()
        self.cap.open(0, cv2.CAP_DSHOW)
        self.cap.set(cv2.CAP_PROP_EXPOSURE,self.exposure)
        for i in range (0,5):
            frame= self.cap.read()
       
        ret,frame = self.cap.read()
        #self.frame=cv2.resize(frame,(912,513))
        self.frame=cv2.resize(frame,(640,480))
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        
        cv2.imwrite("gray/tmp"+str(21)+".jpg",self.frame)
        
        
    def screw_det(self):
        self.img_a=cv2.imread("gray/tmp10.jpg",0)
        self.img_r=cv2.imread('gray/tmp21.jpg',0)
        #img_1=cv2.resize(img_a,(640,480))
        #img_2=cv2.resize(img_r,(640,480))
        iscore=[99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9,99.9]
        if self.current_image_id==1:
            self.no_of_roi=3
            self.roi1_a=self.img_a[107:134,225:253]
            self.roi1_r=self.img_r[107:134,225:253]
            
            self.roi2_a=self.img_a[271:301,339:368]
            self.roi2_r=self.img_r[271:301,339:368]
            
            self.roi3_a=self.img_a[72:96,441:469]
            self.roi3_r=self.img_r[72:96,441:469]

        if self.current_image_id==2:
            self.no_of_roi=2
            self.roi1_a=self.img_a[300:332,163:195]
            self.roi1_r=self.img_r[300:332,163:195]
            
            self.roi2_a=self.img_a[241:269,167:200]
            self.roi2_r=self.img_r[241:269,167:200]

        if self.current_image_id==3: 
            self. no_of_roi=2
            self.roi1_a=self.img_a[310:341,394:428]
            self.roi1_r=self.img_r[310:341,394:428]
            
            self.roi2_a=self.img_a[138:167,395:422]
            self.roi2_r=self.img_r[138:167,395:422]
            
        if self.current_image_id==4:  
            self.no_of_roi=2
            self.roi1_a=self.img_a[127:152,182:208]
            self.roi1_r=self.img_r[127:152,182:208]
            
            self.roi2_a=self.img_a[169:197,447:477]
            self.roi2_r=self.img_r[169:197,447:477]
            
        if self.current_image_id==5:   
            self.no_of_roi=1
            self.roi1_a=self.img_a[225:253,460:492]
            self.roi1_r=self.img_r[225:253,460:492]
            
        if self.current_image_id==6:  
            self.no_of_roi=2
            self.roi1_a=self.img_a[156:182,123:150]
            self.roi1_r=self.img_r[156:182,123:150]
            
            self.roi2_a=self.img_a[251:278,395:424]
            self.roi2_r=self.img_r[251:278,395:424]

        if self.current_image_id==7:
            self.no_of_roi=13
            self.roi1_a=self.img_a[332:359,111:142]
            self.roi1_r=self.img_r[332:359,111:142]
            
            self.roi2_a=self.img_a[211:241,125:152]
            self.roi2_r=self.img_r[211:241,125:152]
            
            self.roi3_a=self.img_a[103:129,138:165]
            self.roi3_r=self.img_r[103:129,138:165]
            
            self.roi4_a=self.img_a[271:302,215:243]
            self.roi4_r=self.img_r[271:302,215:243]
            
            self.roi5_a=self.img_a[156:181,228:255]
            self.roi5_r=self.img_r[156:181,228:255]
            
            self.roi6_a=self.img_a[74:101,261:287]
            self.roi6_r=self.img_r[74:101,261:287]
            
            self.roi7_a=self.img_a[73:96,314:337]
            self.roi7_r=self.img_r[73:96,314:337]
            
            self.roi8_a=self.img_a[335:360,338:365]
            self.roi8_r=self.img_r[335:360,338:365]
            
            self.roi9_a=self.img_a[216:242,338:365]
            self.roi9_r=self.img_r[216:242,338:365]
            
            self.roi10_a=self.img_a[107:131,346:371]
            self.roi10_r=self.img_r[107:131,346:371]
            
            self.roi11_a=self.img_a[92:113,394:415]
            self.roi11_r=self.img_r[92:113,394:415]
            
            self.roi12_a=self.img_a[287:317,465:485]
            self.roi12_r=self.img_r[287:317,465:485]
            
            self.roi13_a=self.img_a[121:142,447:473]
            self.roi13_r=self.img_r[121:142,447:473]
               

            

        print("well done")

        if (self.no_of_roi >0):
            (self.score1, self.diff) = structural_similarity(self.roi1_a,self.roi1_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score1 * 100))
            iscore[1]= abs(round(self.score1 * 10000))
        #  print(iscore[1])

        if (self.no_of_roi > 1):
            (self.score2, self.diff) = structural_similarity(self.roi2_a,self.roi2_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score2 * 100))
            iscore[2]=abs(round(self.score2 * 10000))
        #   print(iscore[2])
            
        if (self.no_of_roi > 2):
            (self.score3, self.diff) = structural_similarity(self.roi3_a,self.roi3_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score3 * 100))
            iscore[3]=abs(round(self.score3 * 10000))
        #   print(iscore[3]) 

        if (self.no_of_roi > 3):
            (self.score4, self.diff) = structural_similarity(self.roi4_a,self.roi4_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score4 * 100))
            iscore[4]=abs(round(self.score4 * 10000))
        #   print(iscore[4]) 

        if (self.no_of_roi > 4):
            (self.score5, self.diff) = structural_similarity(self.roi5_a,self.roi5_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score5 * 100))
            iscore[5]=abs(round(self.score5 * 10000))
        #   print(iscore[5]) 

        if (self.no_of_roi > 5):
            (self.score6, self.diff) = structural_similarity(self.roi6_a,self.roi6_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score6 * 100))
            iscore[6]=abs(round(self.score6 * 10000))
        #   print(iscore[6]) 

        if (self.no_of_roi > 6):
            (self.score7, self.diff) = structural_similarity(self.roi7_a,self.roi7_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score7 * 100))
            iscore[7]=abs(round(self.score7 * 10000))
        #   print(iscore[7]) 

        if (self.no_of_roi > 7):
            (self.score8, self.diff) = structural_similarity(self.roi8_a,self.roi8_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score8 * 100))
            iscore[8]=abs(round(self.score8 * 10000))
        #   print(iscore[8]) 

        if (self.no_of_roi > 8):
            (self.score9, self.diff) = structural_similarity(self.roi9_a,self.roi9_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score9 * 100))
            iscore[9]=abs(round(self.score9 * 10000))
        #   print(iscore[9]) 

        if (self.no_of_roi > 9):
            (self.score10, diff) = structural_similarity(self.roi10_a,self.roi10_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score10 * 100))
            iscore[10]=abs(round(self.score10 * 10000))
        #   print(iscore[10]) 

        if (self.no_of_roi > 10):
            (self.score11, self.diff) = structural_similarity(self.roi11_a,self.roi11_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score11 * 100))
            iscore[11]=abs(round(self.score11 * 10000))
        #   print(iscore[11]) 

        if (self.no_of_roi >11):
            (self.score12, self.diff) = structural_similarity(self.roi12_a,self.roi12_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score12 * 100))
            iscore[12]=abs(round(self.score12 * 10000))
        #   print(iscore[12]) 


        if (self.no_of_roi > 12):
            (self.score13, self.diff) = structural_similarity(self.roi13_a,self.roi13_r, full=True)
            print("Image Similarity: {:.4f}%".format(self.score13 * 100))
            iscore[13]=abs(round(self.score13 * 10000))
        #   print(iscore[13]) 

        self.count=0;
        #self.failedRoi=0;
        while (self.count < self.no_of_roi):
               self.count= self.count+1
               if iscore[self.count] < 4000:
                    self.failedRoi= 0
                    #self.result=1
                    #return self.result
               else: 
         #          print("executed")
                   self.failedRoi= self.count
                   #self.result=0
                   #return self.result
                   break
    
        #return self.result              
        return self.failedRoi
        
        

     

        
        self.cap.release()
        cv2.waitKey(0)
        cv2.destroyAllWindows()
   
                    
         
          
        
if __name__ == '__main__':
    
    file=pathlib.Path("data.xlsx")
    if file.exists ():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet["A1"]="SERIAL NUMBER"
        sheet["B1"]="STEP NUMBER"
        sheet["C1"]="DATE & TIME"
        sheet["D1"]="EXPECTED CYCLE TIME"
        sheet["E1"]="CYCLE TIME"
        sheet["F1"]="RESULT"
        sheet["G1"]="FAILED SCREW"
        file.save("data.xlsx")
    app = App()
    
    app.mainloop()# -*- coding: utf-8 -*-
   
